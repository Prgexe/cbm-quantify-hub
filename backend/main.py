from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import io
import re
import zipfile
from collections import Counter, OrderedDict
from typing import List, Dict, Tuple

app = FastAPI(title="CBMERJ Almoxarifado API")

import sys, logging
logging.basicConfig(level=logging.INFO)
_log = logging.getLogger(__name__)
_log.info(f"Python: {sys.version}")
_log.info(f"openpyxl: {openpyxl.__version__}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def repair_xlsx(data: bytes) -> bytes:
    buf = io.BytesIO(data)
    zin = zipfile.ZipFile(buf, "r")
    SKIP = {"docProps/custom.xml"}
    out = io.BytesIO()
    zout = zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED)
    for item in zin.infolist():
        if item.filename in SKIP:
            continue
        raw = zin.read(item.filename)
        if item.filename.endswith(".xml") or item.filename.endswith(".rels"):
            raw_str = raw.decode("utf-8", errors="replace")
            raw_str = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", raw_str)
            if item.filename == "[Content_Types].xml":
                raw_str = re.sub(r'<Override[^>]*custom\.xml[^/]*/>', '', raw_str)
            raw = raw_str.encode("utf-8")
        zout.writestr(item, raw)
    zout.close()
    out.seek(0)
    return out.read()


def load_workbook_safe(data: bytes, read_only: bool = False, data_only: bool = False):
    try:
        return openpyxl.load_workbook(io.BytesIO(data), read_only=read_only, data_only=data_only)
    except Exception as first_err:
        try:
            repaired = repair_xlsx(data)
            return openpyxl.load_workbook(io.BytesIO(repaired), read_only=read_only, data_only=data_only)
        except Exception:
            raise HTTPException(400, f"Nao foi possivel ler o arquivo Excel. Detalhe: {first_err}")


# ── Helpers ──────────────────────────────────────────────────────────────────

def make_fill(hex_color: str) -> PatternFill:
    rgb = hex_color[-6:] if len(hex_color) > 6 else hex_color
    return PatternFill(fill_type="solid", fgColor=rgb)

def get_color(ws, row_idx: int) -> str:
    try:
        fill = ws.cell(row=row_idx, column=1).fill
        if fill.fill_type is None or fill.fill_type == "none":
            return "FFFFFF"
        rgb = fill.fgColor.rgb
        if rgb in ("00000000", "000000"):
            return "FFFFFF"
        return rgb[-6:] if len(rgb) > 6 else rgb
    except:
        return "FFFFFF"

POSTO_MAP = {
    "CEL": "CORONEL", "COR": "CORONEL",
    "TEN CEL": "TENENTE CORONEL", "TENCEL": "TENENTE CORONEL",
    "TEM CEL": "TENENTE CORONEL", "TEMCEL": "TENENTE CORONEL",
    "TC": "TENENTE CORONEL", "T CEL": "TENENTE CORONEL",
    "TENENTE-CORONEL": "TENENTE CORONEL",
    "MAJ": "MAJOR",
    "CAP": "CAPITAO",
    "TEN": "TENENTE",
    "1 TEN": "1º TENENTE", "1TEN": "1º TENENTE", "1º TEN": "1º TENENTE",
    "2 TEN": "2º TENENTE", "2TEN": "2º TENENTE", "2º TEN": "2º TENENTE",
    "SUBTEN": "SUBTENENTE", "ST": "SUBTENENTE", "SUB TEN": "SUBTENENTE", "SUB-TEN": "SUBTENENTE",
    "1 SGT": "1º SARGENTO", "1SGT": "1º SARGENTO", "1º SGT": "1º SARGENTO", "1° SGT": "1º SARGENTO",
    "2 SGT": "2º SARGENTO", "2SGT": "2º SARGENTO", "2º SGT": "2º SARGENTO", "2° SGT": "2º SARGENTO",
    "3 SGT": "3º SARGENTO", "3SGT": "3º SARGENTO", "3º SGT": "3º SARGENTO", "3° SGT": "3º SARGENTO",
    "CB": "CABO", "SD": "SOLDADO", "SOL": "SOLDADO",
}

def normalizar_posto(posto):
    if not posto:
        return posto
    key = str(posto).strip().upper()
    if key in POSTO_MAP:
        return POSTO_MAP[key]
    key_limpa = key.replace("°", "").replace("º", "")
    if key_limpa in POSTO_MAP:
        return POSTO_MAP[key_limpa]
    key_no_spaces = " ".join(key_limpa.split())
    return POSTO_MAP.get(key_no_spaces, posto)


def normalizar_area(area: str, areas_consolidada: list) -> str:
    if not area or not areas_consolidada:
        return area
    normalized_upper = area.strip().upper()
    for a in areas_consolidada:
        if a.strip().upper() == normalized_upper:
            return a
    for a in areas_consolidada:
        if a.strip().upper().startswith(normalized_upper):
            return a
    for a in areas_consolidada:
        if normalized_upper in a.strip().upper():
            return a
    return area.strip()


def norm(s) -> str:
    return str(s or "").strip().upper()

# Colunas fixas/estruturais que não são materiais
FIXED_COLUMNS = {"QTD", "AREA", "ÁREA", "UNIDADE", "UNIDADE (FINAL)", "POSTO/GRAD",
                  "QUADRO", "NOME COMPLETO", "RG", "TAMANHOS"}

def is_material_column(col_name: str) -> bool:
    """Retorna True se a coluna é de material (não é estrutural)."""
    n = norm(col_name)
    if not n:
        return False
    return n not in FIXED_COLUMNS


# ── Leitura DINÂMICA de planilha individual ──────────────────────────────────

def find_header_rows(raw: list) -> Tuple[int, Dict[int, str], bool]:
    """
    Detecta linhas de cabeçalho e retorna:
    - data_start: índice da primeira linha de dados
    - col_map: {col_index: col_name}
    - is_format_b: se é formato B (cabeçalho em 2 linhas)
    """
    is_format_b = len(raw) > 1 and any(norm(c) == "QTD" for c in (raw[1] or []))

    col_map = {}

    if is_format_b:
        # Formato B: row[1] tem colunas fixas, row[2] tem nomes de materiais
        # row[2] tem PRIORIDADE sobre row[1] para colunas de material
        for i, c in enumerate(raw[1] or []):
            n = norm(c)
            if n:
                col_map[i] = n
        for i, c in enumerate(raw[2] or []):
            n = norm(c)
            # Sobrescreve SEMPRE que row[2] tem valor — nomes de material
            # têm prioridade sobre 'TAMANHOS' genérico de row[1]
            if n:
                col_map[i] = n
        return 3, col_map, True
    else:
        # Formato A: row[2] tem todas as colunas
        for i, c in enumerate(raw[2] or []):
            n = norm(c)
            if n:
                col_map[i] = n
        return 3, col_map, False


def normalize_col_name(n: str) -> str:
    """Normaliza nomes de colunas para chaves padronizadas."""
    n = n.strip().upper()
    if n in ("ÁREA", "AREA"):
        return "AREA"
    if n in ("UNIDADE (FINAL)", "UNIDADE"):
        return "UNIDADE"
    if n == "POSTO/GRAD":
        return "POSTO_GRAD"
    if n == "NOME COMPLETO":
        return "NOME_COMPLETO"
    # Para materiais e outros campos, mantém o nome original normalizado
    return n


def read_individual_sheet(ws) -> Tuple[List[dict], List[str]]:
    """
    Lê uma aba da planilha individual.
    Retorna (records, material_columns) onde material_columns são os nomes
    das colunas de material encontradas.
    """
    raw = list(ws.iter_rows(values_only=True))
    if len(raw) < 4:
        return [], []

    data_start, col_map, is_format_b = find_header_rows(raw)

    # Identifica colunas de material
    material_cols = []  # lista de (col_index, col_name_original)
    fixed_map = {}      # col_index -> normalized_key

    for idx, col_name in col_map.items():
        if is_material_column(col_name):
            material_cols.append((idx, col_name))
        else:
            fixed_map[idx] = normalize_col_name(col_name)

    material_names = [name for _, name in material_cols]

    records = []
    for row in raw[data_start:]:
        if not row or all(str(v or "").strip() == "" for v in row):
            continue

        rec = {"QTD": "", "AREA": "", "UNIDADE": "", "POSTO_GRAD": "",
               "QUADRO": "", "NOME_COMPLETO": "", "RG": ""}

        # Campos fixos
        for idx, key in fixed_map.items():
            if idx < len(row):
                rec[key] = str(row[idx] or "").strip()

        # Campos de material (dinâmicos)
        for idx, col_name in material_cols:
            if idx < len(row):
                rec[col_name] = str(row[idx] or "").strip()

        rec["POSTO_GRAD"] = normalizar_posto(rec.get("POSTO_GRAD", "")) or rec.get("POSTO_GRAD", "")

        if rec.get("NOME_COMPLETO"):
            records.append(rec)

    _log.info(f"[read_individual] Colunas de material detectadas: {material_names}")
    _log.info(f"[read_individual] {len(records)} registros lidos")

    return records, material_names


# ── Leitura do cabeçalho da consolidada ──────────────────────────────────────

def read_consolidada_header(ws) -> Tuple[int, Dict[int, str], Dict[str, int]]:
    """
    Lê o cabeçalho da aba da consolidada.
    Detecta automaticamente a linha do cabeçalho procurando por
    colunas-chave como ÁREA, UNIDADE, NOME COMPLETO ou QTD.
    """
    header_row = 1  # default
    KEY_HEADERS = {"QTD", "AREA", "ÁREA", "UNIDADE", "UNIDADE (FINAL)", "NOME COMPLETO", "NOME_COMPLETO"}
    for ri in range(1, 10):
        row_vals = [norm(ws.cell(row=ri, column=c).value) for c in range(1, min(ws.max_column+1, 12))]
        matches = sum(1 for v in row_vals if v in KEY_HEADERS)
        if matches >= 2:  # linha com pelo menos 2 colunas conhecidas = cabeçalho
            header_row = ri
            break

    col_map = {}
    name_to_col = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            n = norm(val)
            col_map[col] = n
            name_to_col[n] = col

    # Também verifica a linha seguinte para cabeçalhos de 2 linhas
    for col in range(1, ws.max_column + 1):
        if col not in col_map or not col_map[col]:
            val = ws.cell(row=header_row + 1, column=col).value
            if val:
                n = norm(val)
                col_map[col] = n
                if n not in name_to_col:
                    name_to_col[n] = col

    return header_row, col_map, name_to_col


def fuzzy_match_column(material_name: str, consolidada_cols: Dict[str, int]) -> int | None:
    """
    Tenta encontrar a coluna correspondente na consolidada para um material.
    Usa matching exato primeiro, depois parcial.
    """
    n = norm(material_name)

    # Exato
    if n in consolidada_cols:
        return consolidada_cols[n]

    # Parcial: verifica se o nome do material está contido em alguma coluna
    for col_name, col_idx in consolidada_cols.items():
        if n in col_name or col_name in n:
            return col_idx

    # Normaliza removendo acentos comuns e pontuação
    n_clean = n.replace(".", "").replace("-", " ").replace("_", " ")
    for col_name, col_idx in consolidada_cols.items():
        col_clean = col_name.replace(".", "").replace("-", " ").replace("_", " ")
        if n_clean in col_clean or col_clean in n_clean:
            return col_idx

    return None


# ── Endpoint: info ───────────────────────────────────────────────────────────

IGNORED_SHEETS = {"Resumo Geral", "Contagem", "Acrescentar1"}

@app.post("/info")
async def get_info(file: UploadFile = File(...)):
    data = await file.read()
    wb = load_workbook_safe(data, read_only=True, data_only=True)

    sheets = []
    for name in wb.sheetnames:
        if name in IGNORED_SHEETS:
            continue
        ws = wb[name]
        unidades = set()
        row_count = 0
        data_start = 2
        unidade_col_idx = 1  # fallback: índice 1 = segunda coluna (UNIDADE)
        # Detecta linha do cabeçalho procurando por UNIDADE ou QTD
        for ri, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
            if not row: continue
            row_upper = [str(v or '').strip().upper() for v in row]
            # Cabeçalho: linha que contém UNIDADE ou QTD
            if 'UNIDADE' in row_upper or 'UNIDADE (FINAL)' in row_upper or 'QTD' in row_upper:
                data_start = ri + 1
                for ci, val in enumerate(row_upper):
                    if val in ('UNIDADE', 'UNIDADE (FINAL)'):
                        unidade_col_idx = ci
                        break
                break
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            if row and any(v is not None for v in row):
                row_count += 1
                if unidade_col_idx < len(row) and row[unidade_col_idx]:
                    unidades.add(str(row[unidade_col_idx]).strip())
        sheets.append({
            "name": name,
            "rows": row_count + 3,
            "unidades": sorted(unidades - {""})
        })
    wb.close()
    return {"sheets": sheets}


# ── Endpoint: merge ──────────────────────────────────────────────────────────

@app.post("/merge")
async def merge(
    consolidada: UploadFile = File(...),
    individual:  UploadFile = File(...),
    aba_destino: str = Form(""),
    inserir_antes_de: str = Form(""),
    inserir_modo: str = Form("depois"),
):
    consolidada_bytes = await consolidada.read()
    individual_bytes  = await individual.read()

    wb_cons = load_workbook_safe(consolidada_bytes)
    wb_ind  = load_workbook_safe(individual_bytes)

    # Lê todos os militares da planilha individual (com colunas dinâmicas)
    novos = []
    all_material_cols = []
    for sheet_name in wb_ind.sheetnames:
        ws_ind = wb_ind[sheet_name]
        records, mat_cols = read_individual_sheet(ws_ind)
        novos.extend(records)
        for mc in mat_cols:
            if mc not in all_material_cols:
                all_material_cols.append(mc)

    if not novos:
        raise HTTPException(400, "Nenhum militar encontrado na planilha individual.")

    _log.info(f"[merge] {len(novos)} militares, materiais: {all_material_cols}")

    aba_destino = aba_destino.strip()
    inserir_antes_de = inserir_antes_de.strip()
    if inserir_antes_de == "__fim__":
        inserir_antes_de = ""

    if aba_destino not in wb_cons.sheetnames:
        raise HTTPException(400, f"Aba '{aba_destino}' nao encontrada na planilha consolidada.")

    ws_dest = wb_cons[aba_destino]

    # Lê o cabeçalho da consolidada para mapear colunas
    header_row, cons_col_map, cons_name_to_col = read_consolidada_header(ws_dest)
    data_start_row = header_row + 1

    _log.info(f"[merge] Cabeçalho consolidada (linha {header_row}): {cons_col_map}")

    # Mapeia as colunas fixas da consolidada
    fixed_col_mapping = {}
    for key in ["QTD", "AREA", "ÁREA", "UNIDADE", "UNIDADE (FINAL)", "POSTO/GRAD",
                "QUADRO", "NOME COMPLETO", "RG"]:
        if key in cons_name_to_col:
            fixed_col_mapping[normalize_col_name(key)] = cons_name_to_col[key]

    # Mapeia as colunas de material: individual -> consolidada col index
    material_col_mapping = {}  # material_name -> col_index na consolidada
    for mat_name in all_material_cols:
        col_idx = fuzzy_match_column(mat_name, cons_name_to_col)
        if col_idx:
            material_col_mapping[mat_name] = col_idx
            _log.info(f"[merge] Material '{mat_name}' -> coluna {col_idx}")
        else:
            _log.warning(f"[merge] Material '{mat_name}' NÃO encontrado na consolidada!")

    # Coleta areas reais da aba destino
    areas_consolidada = []
    for row in ws_dest.iter_rows(min_row=data_start_row, values_only=True):
        if row and row[1]:
            val = str(row[1]).strip()
            if val and val not in areas_consolidada:
                areas_consolidada.append(val)

    # Normaliza area e posto
    for mil in novos:
        mil["AREA"] = normalizar_area(mil.get("AREA", ""), areas_consolidada)
        mil["UNIDADE"] = normalizar_area(mil.get("UNIDADE", ""), areas_consolidada)
        mil["POSTO_GRAD"] = normalizar_posto(mil.get("POSTO_GRAD", "")) or mil.get("POSTO_GRAD", "")

    # Encontra linha de inserção
    insert_row = None
    if inserir_antes_de:
        unidade_col = cons_name_to_col.get("UNIDADE (FINAL)") or cons_name_to_col.get("UNIDADE") or 3
        first_row = None
        last_row  = None
        for i in range(data_start_row, ws_dest.max_row + 1):
            val = ws_dest.cell(row=i, column=unidade_col).value
            if val and str(val).strip() == inserir_antes_de:
                if first_row is None: first_row = i
                last_row = i
        if inserir_modo == "depois" and last_row is not None:
            insert_row = last_row + 1   # insere após a última linha da unidade
        elif first_row is not None:
            insert_row = first_row      # insere antes da primeira linha da unidade
    if insert_row is None:
        insert_row = ws_dest.max_row + 1

    n = len(novos)

    cores_originais = {}
    for r in range(data_start_row, ws_dest.max_row + 1):
        cores_originais[r] = get_color(ws_dest, r)

    ref = list(ws_dest[insert_row]) if insert_row <= ws_dest.max_row else []

    ws_dest.insert_rows(insert_row, amount=n)

    COR_AZUL = "DAEEF3"
    COR_BRANCO = "FFFFFF"

    for i, mil in enumerate(novos):
        rn = insert_row + i
        cor = COR_AZUL if i % 2 == 0 else COR_BRANCO
        fill = make_fill(cor)

        # Escreve campos fixos
        fixed_values = {
            "QTD": i + 1,
            "AREA": mil.get("AREA", ""),
            "UNIDADE": mil.get("UNIDADE", ""),
            "POSTO_GRAD": mil.get("POSTO_GRAD", ""),
            "QUADRO": mil.get("QUADRO", ""),
            "NOME_COMPLETO": mil.get("NOME_COMPLETO", ""),
            "RG": mil.get("RG", ""),
        }

        for key, val in fixed_values.items():
            col_idx = fixed_col_mapping.get(key)
            if col_idx:
                cell = ws_dest.cell(row=rn, column=col_idx, value=val)
                cell.fill = fill
                if col_idx - 1 < len(ref) and ref[col_idx - 1].font:
                    cell.font = copy(ref[col_idx - 1].font)
                if col_idx - 1 < len(ref) and ref[col_idx - 1].alignment:
                    cell.alignment = copy(ref[col_idx - 1].alignment)
                if col_idx - 1 < len(ref) and ref[col_idx - 1].border:
                    cell.border = copy(ref[col_idx - 1].border)

        # Escreve campos de material (dinâmicos)
        for mat_name, col_idx in material_col_mapping.items():
            val = mil.get(mat_name, "")
            cell = ws_dest.cell(row=rn, column=col_idx, value=val)
            cell.fill = fill
            if col_idx - 1 < len(ref) and ref[col_idx - 1].font:
                cell.font = copy(ref[col_idx - 1].font)
            if col_idx - 1 < len(ref) and ref[col_idx - 1].alignment:
                cell.alignment = copy(ref[col_idx - 1].alignment)
            if col_idx - 1 < len(ref) and ref[col_idx - 1].border:
                cell.border = copy(ref[col_idx - 1].border)

        # Preenche colunas restantes com fill (para manter visual consistente)
        for col in range(1, ws_dest.max_column + 1):
            cell = ws_dest.cell(row=rn, column=col)
            if cell.value is None:
                cell.fill = fill

    # Restaura cores originais das linhas deslocadas
    for old_row, cor in cores_originais.items():
        new_row = old_row + n
        if new_row <= ws_dest.max_row:
            fill = make_fill(cor)
            for col in range(1, ws_dest.max_column + 1):
                ws_dest.cell(row=new_row, column=col).fill = fill

    # ── Atualiza Consolidado Geral (se existir) ─────────────────────────────────
    if "Consolidado Geral" in wb_cons.sheetnames:
        ws_cg = wb_cons["Consolidado Geral"]

        # Lê cabeçalho do Consolidado Geral
        _, cg_col_map, cg_name_to_col = read_consolidada_header(ws_cg)
        cg_data_start = list(cg_col_map.keys())[0] if cg_col_map else 1
        # Detecta header_row do Consolidado Geral
        for ri in range(1, 10):
            row_vals = [norm(ws_cg.cell(ri, c).value) for c in range(1, min(ws_cg.max_column+1, 12))]
            if sum(1 for v in row_vals if v in {"AREA","ÁREA","UNIDADE","NOME COMPLETO","QTD"}) >= 2:
                cg_header_row = ri
                break
        else:
            cg_header_row = 1

        _, cg_col_map_full, cg_name_to_col_full = read_consolidada_header(ws_cg)

        # Mapeamento de colunas fixas no CG
        cg_fixed = {}
        for key in ["AREA","ÁREA","UNIDADE","UNIDADE (FINAL)","POSTO/GRAD","QUADRO","NOME COMPLETO","RG"]:
            if key in cg_name_to_col_full:
                cg_fixed[normalize_col_name(key)] = cg_name_to_col_full[key]

        # Mapeamento de materiais no CG
        cg_mat_cols = {}
        for mat in all_material_cols:
            col_idx = fuzzy_match_column(mat, cg_name_to_col_full)
            if col_idx:
                cg_mat_cols[mat] = col_idx

        # Adiciona militares ao Consolidado Geral no final
        for mil in novos:
            new_row = ws_cg.max_row + 1
            for key, col_idx in cg_fixed.items():
                ws_cg.cell(new_row, col_idx).value = mil.get(key, "")
            for mat, col_idx in cg_mat_cols.items():
                ws_cg.cell(new_row, col_idx).value = mil.get(mat, "")

        _log.info(f"[merge] {len(novos)} militares adicionados ao Consolidado Geral")

    # ── Atualiza aba Contagem (se existir) ────────────────────────────────────
    if "Contagem" in wb_cons.sheetnames:
        ws_cont = wb_cons["Contagem"]

        # Descobre materiais únicos da Contagem (coluna B, exceto cabeçalho "Material")
        materiais_contagem = []
        for i in range(1, ws_cont.max_row + 1):
            mat = str(ws_cont.cell(i, 2).value or "").strip()
            if mat and mat != "Material" and mat not in materiais_contagem:
                materiais_contagem.append(mat)

        # Unidades já existentes na Contagem (col A)
        unidades_existentes = set()
        for i in range(1, ws_cont.max_row + 1):
            val = ws_cont.cell(i, 1).value
            if val:
                unidades_existentes.add(str(val).strip())

        unidades_novas = set(m.get("UNIDADE", "") for m in novos if m.get("UNIDADE"))
        unidades_para_adicionar = sorted(unidades_novas - unidades_existentes)

        if unidades_para_adicionar and materiais_contagem:
            # Encontra linha de referência para estilos (primeira linha de dados)
            ref_row_idx = 2
            ref_cont = list(ws_cont[ref_row_idx]) if ref_row_idx <= ws_cont.max_row else []

            # Detecta mapeamento de material -> coluna do Consolidado Geral
            # Lê do cabeçalho da aba Consolidado Geral
            mat_to_cg_col = {}
            if "Consolidado Geral" in wb_cons.sheetnames:
                ws_cg = wb_cons["Consolidado Geral"]
                cg_header = {
                    str(ws_cg.cell(1, c).value or "").strip().upper(): c
                    for c in range(1, ws_cg.max_column + 1)
                }
                for mat in materiais_contagem:
                    mat_upper = mat.strip().upper()
                    if mat_upper in cg_header:
                        mat_to_cg_col[mat] = get_column_letter(cg_header[mat_upper])
                    else:
                        # fuzzy: verifica se nome do material está contido
                        for h, ci in cg_header.items():
                            if mat_upper in h or h in mat_upper:
                                mat_to_cg_col[mat] = get_column_letter(ci)
                                break
            last_cg_row = ws_cg.max_row if "Consolidado Geral" in wb_cons.sheetnames else 2000

            # Descobre tamanhos da contagem (colunas entre Material e TOTAL)
            # Lê do cabeçalho da linha 1 da Contagem
            sizes_cols = {}  # tamanho -> col_letter
            for c in range(3, ws_cont.max_column + 1):
                h = str(ws_cont.cell(1, c).value or "").strip()
                if h and h != "TOTAL":
                    sizes_cols[h] = get_column_letter(c)
                elif h == "TOTAL":
                    total_col_letter = get_column_letter(c)
                    break

            # Determina onde inserir na Contagem
            insert_cont = ws_cont.max_row + 1
            if inserir_antes_de:
                first_ref = last_ref = None
                for i in range(1, ws_cont.max_row + 1):
                    val = ws_cont.cell(i, 1).value
                    if val and str(val).strip() == inserir_antes_de:
                        if first_ref is None: first_ref = i
                        last_ref = i
                if inserir_modo == "depois" and last_ref:
                    insert_cont = last_ref + 1
                elif first_ref:
                    insert_cont = first_ref

            n_mats = len(materiais_contagem)
            rows_inseridos = len(unidades_para_adicionar) * n_mats
            ws_cont.insert_rows(insert_cont, amount=rows_inseridos)

            current_row = insert_cont
            for unidade in unidades_para_adicionar:
                for mi, material in enumerate(materiais_contagem):
                    rn = current_row + mi
                    # Col 1: Unidade (só na primeira linha do bloco)
                    ws_cont.cell(rn, 1).value = unidade if mi == 0 else None
                    # Col 2: Material
                    ws_cont.cell(rn, 2).value = material
                    # Colunas de tamanho: fórmulas COUNTIFS -> Consolidado Geral
                    cg_col = mat_to_cg_col.get(material, "")
                    for sz, sz_col_letter in sizes_cols.items():
                        sz_col = column_index_from_string(sz_col_letter)
                        cell = ws_cont.cell(rn, sz_col)
                        if cg_col:
                            cell.value = (
                                f"=COUNTIFS('Consolidado Geral'!B$2:B${last_cg_row},"
                                f"\"{unidade}\",'Consolidado Geral'!{cg_col}$2:{cg_col}${last_cg_row},"
                                f"\"{sz}\")"
                            )
                        else:
                            cell.value = "--"
                    # Coluna TOTAL
                    last_sz_col = max(column_index_from_string(c) for c in sizes_cols.values())
                    first_sz_col = min(column_index_from_string(c) for c in sizes_cols.values())
                    tot_col = last_sz_col + 1
                    ws_cont.cell(rn, tot_col).value = (
                        f"=SUM({get_column_letter(first_sz_col)}{rn}:{get_column_letter(last_sz_col)}{rn})"
                    )
                    # Copia estilos da linha de referência
                    for col in range(1, ws_cont.max_column + 1):
                        if col <= len(ref_cont) and ref_cont[col-1].has_style:
                            c = ws_cont.cell(rn, col)
                            c.font      = copy(ref_cont[col-1].font)
                            c.alignment = copy(ref_cont[col-1].alignment)
                            c.border    = copy(ref_cont[col-1].border)
                            c.fill      = copy(ref_cont[col-1].fill)
                current_row += n_mats

    # ── Atualiza Resumo Geral (se existir) ────────────────────────────────────
    if "Resumo Geral" in wb_cons.sheetnames and "Contagem" in wb_cons.sheetnames:
        ws_cont = wb_cons["Contagem"]
        ws_res  = wb_cons["Resumo Geral"]

        # Coleta todos os tipos de material da aba Contagem (coluna B)
        materiais_contagem = {}  # material_name -> [row_indices]
        for i in range(1, ws_cont.max_row + 1):
            mat = ws_cont.cell(row=i, column=2).value
            if mat:
                mat_str = str(mat).strip()
                if mat_str:
                    if mat_str not in materiais_contagem:
                        materiais_contagem[mat_str] = []
                    materiais_contagem[mat_str].append(i)

        # Para cada material no Resumo Geral, atualiza as fórmulas SUM
        for res_row in range(1, ws_res.max_row + 1):
            mat_res = ws_res.cell(row=res_row, column=1).value
            if not mat_res:
                continue
            mat_res_str = str(mat_res).strip()
            if mat_res_str in materiais_contagem:
                rows = materiais_contagem[mat_res_str]
                # Atualiza cada coluna de tamanho
                for col in range(2, ws_res.max_column + 1):
                    existing = ws_res.cell(row=res_row, column=col).value
                    if existing and str(existing).startswith("="):
                        # Reconstrói a fórmula SUM com as linhas corretas
                        # Detecta qual coluna da Contagem está sendo referenciada
                        match = re.match(r"=.*Contagem!([A-Z]+)\d+", str(existing))
                        if match:
                            cont_col = match.group(1)
                            formula = "=" + "+".join(f"Contagem!{cont_col}{r}" for r in rows)
                            ws_res.cell(row=res_row, column=col).value = formula

    # ── Retorna arquivo ───────────────────────────────────────────────────────
    output = io.BytesIO()
    wb_cons.save(output)
    output.seek(0)

    filename = consolidada.filename.replace(".xlsx", "_atualizado.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/health")
def health():
    return {"status": "ok"}
