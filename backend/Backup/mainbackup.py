from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy
import io
import re
from collections import Counter
from typing import List

app = FastAPI(title="CBMERJ Almoxarifado API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Helpers ──────────────────────────────────────────────────────────────────

def make_fill(hex_color: str) -> PatternFill:
    rgb = hex_color[-6:] if len(hex_color) > 6 else hex_color
    return PatternFill(fill_type="solid", fgColor=rgb)

def get_color(ws, row_idx: int) -> str:
    try:
        rgb = ws.cell(row=row_idx, column=1).fill.fgColor.rgb
        return rgb[-6:] if len(rgb) > 6 else rgb
    except:
        return "FFFFFF"

POSTO_MAP = {
    "CEL": "CORONEL", "TEM CEL": "TENENTE CORONEL", "TEN CEL": "TENENTE CORONEL",
    "TENENTE-CORONEL": "TENENTE CORONEL", "MAJ": "MAJOR", "CAP": "CAPITÃO",
    "TEN": "TENENTE", "1º TEN": "1º TENENTE", "2º TEN": "2º TENENTE",
    "SUBTEN": "SUBTENENTE", "ST": "SUBTENENTE",
    "1º SGT": "1º SARGENTO", "1SGT": "1º SARGENTO",
    "2º SGT": "2º SARGENTO", "2SGT": "2º SARGENTO",
    "3º SGT": "3º SARGENTO", "3SGT": "3º SARGENTO",
    "CB": "CABO", "SD": "SOLDADO", "SOL": "SOLDADO",
}

def normalizar_posto(posto):
    if not posto:
        return posto
    return POSTO_MAP.get(str(posto).strip().upper(), posto)


# ── Leitura de planilha individual ───────────────────────────────────────────

def norm(s) -> str:
    return str(s or "").strip().upper()

def read_individual_sheet(ws) -> List[dict]:
    """Lê aba de planilha individual (Formato B: 3 linhas de cabeçalho)."""
    raw = list(ws.iter_rows(values_only=True))
    if len(raw) < 4:
        return []

    col_map = {}

    def map_cell(n, i):
        if n == "QTD":                                          col_map[i] = "QTD"
        elif n in ("ÁREA", "AREA"):                             col_map[i] = "AREA"
        elif n in ("UNIDADE (FINAL)", "UNIDADE"):               col_map[i] = "UNIDADE"
        elif n == "POSTO/GRAD":                                 col_map[i] = "POSTO_GRAD"
        elif n == "QUADRO":                                     col_map[i] = "QUADRO"
        elif n == "NOME COMPLETO":                              col_map[i] = "NOME_COMPLETO"
        elif n == "RG":                                         col_map[i] = "RG"
        elif n in ("CAMISETA DE GV", "CAMISETA GV"):            col_map[i] = "CAMISETA_GV"
        elif n.startswith("CAMISA U"):                          col_map[i] = "CAMISA_UV"
        elif "SHORT JOHN" in n:                                 col_map[i] = "SHORT_JOHN"

    is_format_b = any(norm(c) == "QTD" for c in (raw[1] or []))
    if is_format_b:
        for i, c in enumerate(raw[1] or []): map_cell(norm(c), i)
        for i, c in enumerate(raw[2] or []): map_cell(norm(c), i)
    else:
        for i, c in enumerate(raw[2] or []): map_cell(norm(c), i)

    records = []
    for row in raw[3:]:
        if not row or all(str(v or "").strip() == "" for v in row):
            continue
        rec = {f: "" for f in ["QTD","AREA","UNIDADE","POSTO_GRAD","QUADRO","NOME_COMPLETO","RG","CAMISETA_GV","CAMISA_UV","SHORT_JOHN"]}
        for idx, val in enumerate(row):
            field = col_map.get(idx)
            if field:
                rec[field] = str(val or "").strip()
        rec["POSTO_GRAD"] = normalizar_posto(rec["POSTO_GRAD"]) or rec["POSTO_GRAD"]
        if rec["NOME_COMPLETO"]:
            records.append(rec)
    return records


# ── Endpoint: info da planilha consolidada ───────────────────────────────────

@app.post("/info")
async def get_info(file: UploadFile = File(...)):
    """Retorna as abas disponíveis e unidades da planilha consolidada."""
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        unidades = set()
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row and row[2]:
                unidades.add(str(row[2]).strip())
        sheets.append({
            "name": name,
            "rows": ws.max_row,
            "unidades": sorted(unidades - {""})
        })
    
    return {"sheets": sheets}


# ── Endpoint: mesclar planilha individual na consolidada ─────────────────────

@app.post("/merge")
async def merge(
    consolidada: UploadFile = File(...),
    individual:  UploadFile = File(...),
    aba_destino: str = Form(""),
    inserir_antes_de: str = Form(""),
):
    """
    Insere os militares da planilha individual na aba da consolidada,
    antes da unidade especificada, preservando formatação e atualizando
    a aba Contagem e Resumo Geral.
    """
    consolidada_bytes = await consolidada.read()
    individual_bytes  = await individual.read()

    wb_cons = openpyxl.load_workbook(io.BytesIO(consolidada_bytes))
    wb_ind  = openpyxl.load_workbook(io.BytesIO(individual_bytes))

    # Lê todos os militares da planilha individual
    novos = []
    for sheet_name in wb_ind.sheetnames:
        ws_ind = wb_ind[sheet_name]
        novos.extend(read_individual_sheet(ws_ind))

    if not novos:
        raise HTTPException(400, "Nenhum militar encontrado na planilha individual.")

    # Seleciona aba destino
    if aba_destino not in wb_cons.sheetnames:
        raise HTTPException(400, f"Aba '{aba_destino}' não encontrada na planilha consolidada.")

    ws_dest = wb_cons[aba_destino]

    # Encontra linha de inserção
    insert_row = None
    if inserir_antes_de:
        for i in range(4, ws_dest.max_row + 1):
            val = ws_dest.cell(row=i, column=3).value
            if val and str(val).strip() == inserir_antes_de:
                insert_row = i
                break
    if insert_row is None:
        insert_row = ws_dest.max_row + 1

    n = len(novos)

    # Salva cores originais antes de inserir
    cores_originais = {}
    for r in range(4, ws_dest.max_row + 1):
        cores_originais[r] = get_color(ws_dest, r)

    # Linha de referência para estilo
    ref = list(ws_dest[insert_row]) if insert_row <= ws_dest.max_row else []

    # Insere linhas
    ws_dest.insert_rows(insert_row, amount=n)

    COR_AZUL   = "DAEEF3"
    COR_BRANCO = "FFFFFF"

    # Preenche novos militares
    for i, mil in enumerate(novos):
        rn = insert_row + i
        cor = COR_AZUL if i % 2 == 0 else COR_BRANCO
        fill = make_fill(cor)
        values = [
            i + 1,
            mil["AREA"],
            mil["UNIDADE"],
            mil["POSTO_GRAD"],
            mil["QUADRO"],
            mil["NOME_COMPLETO"],
            mil["RG"],
            mil["CAMISETA_GV"],
            mil["CAMISA_UV"],
            mil["SHORT_JOHN"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_dest.cell(row=rn, column=col_idx, value=val)
            cell.fill = fill
            if col_idx <= len(ref) and ref[col_idx-1].font:
                cell.font = copy(ref[col_idx-1].font)
            if col_idx <= len(ref) and ref[col_idx-1].alignment:
                cell.alignment = copy(ref[col_idx-1].alignment)
            if col_idx <= len(ref) and ref[col_idx-1].border:
                cell.border = copy(ref[col_idx-1].border)

    # Reaplica cores originais nas linhas deslocadas
    for old_row, cor in cores_originais.items():
        new_row = old_row + n
        if new_row <= ws_dest.max_row:
            fill = make_fill(cor)
            for col in range(1, ws_dest.max_column + 1):
                ws_dest.cell(row=new_row, column=col).fill = fill

    # ── Atualiza aba Contagem ─────────────────────────────────────────────────
    if "Contagem" in wb_cons.sheetnames:
        ws_cont = wb_cons["Contagem"]

        # Unidades novas que ainda não existem na Contagem
        unidades_novas = set(m["UNIDADE"] for m in novos if m["UNIDADE"])
        unidades_existentes = set()
        for i in range(1, ws_cont.max_row + 1):
            val = ws_cont.cell(row=i, column=1).value
            if val:
                unidades_existentes.add(str(val).strip())

        unidades_para_adicionar = unidades_novas - unidades_existentes

        # Encontra onde inserir na Contagem (antes da unidade inserir_antes_de)
        insert_cont = ws_cont.max_row + 1
        if inserir_antes_de:
            for i in range(1, ws_cont.max_row + 1):
                val = ws_cont.cell(row=i, column=1).value
                if val and str(val).strip() == inserir_antes_de:
                    insert_cont = i
                    break

        sheet_ref = f"'{aba_destino}'"
        materiais = [
            ("Camiseta de GV", "$H", ["PP","P","M","G","GG","XGG","EXG","XL","--"]),
            ("Camisa U.V",     "$I", ["P","M","G","GG","XGG","--"]),
            ("Short John",     "$J", ["P","M","G","GG","XGG","--"]),
        ]
        extras = {
            "$H": [("XG","XG"),("XXG","XXG"),("3G","3G")],
            "$I": [("XG","XG"),("XXG","XXG"),("EXG","EXG"),("3G","3G")],
            "$J": [("XG","XG"),("XXG","XXG"),("EXG","EXG"),("XL","XL"),("5XL","5XL")],
        }

        ref_cont = list(ws_cont[max(1, insert_cont - 1)])
        rows_inseridos = len(unidades_para_adicionar) * 3

        if rows_inseridos > 0:
            ws_cont.insert_rows(insert_cont, amount=rows_inseridos)

        current_row = insert_cont
        for unidade in sorted(unidades_para_adicionar):
            for offset, (material, col_mat, tamanhos) in enumerate(materiais):
                rn = current_row + offset
                ws_cont.cell(row=rn, column=1).value = unidade if offset == 0 else None
                ws_cont.cell(row=rn, column=2).value = material
                for c_idx, tam in enumerate(tamanhos):
                    f = f'=COUNTIFS({sheet_ref}!$C$4:$C$10000,"{unidade}",{sheet_ref}!{col_mat}$4:{col_mat}$10000,"{tam}")'
                    ws_cont.cell(row=rn, column=c_idx + 3).value = f
                # TOTAL col K
                sum_f = f'=SUM(C{rn}:I{rn})'
                for tam, label in extras[col_mat]:
                    sum_f += f'+COUNTIFS({sheet_ref}!$C$4:$C$10000,"{unidade}",{sheet_ref}!{col_mat}$4:{col_mat}$10000,"{tam}")'
                ws_cont.cell(row=rn, column=11).value = sum_f
                # Estilo
                for col in range(1, ws_cont.max_column + 1):
                    cell = ws_cont.cell(row=rn, column=col)
                    if col <= len(ref_cont) and ref_cont[col-1].font:
                        cell.font = copy(ref_cont[col-1].font)
                    if col <= len(ref_cont) and ref_cont[col-1].alignment:
                        cell.alignment = copy(ref_cont[col-1].alignment)
                    if col <= len(ref_cont) and ref_cont[col-1].border:
                        cell.border = copy(ref_cont[col-1].border)
            current_row += 3

        # Corrige fórmulas SUM que apontam para linha errada após insert_rows
        for r in range(1, ws_cont.max_row + 1):
            cell_k = ws_cont.cell(row=r, column=11)
            val = str(cell_k.value or "")
            if not val.startswith("=SUM"):
                continue
            m = re.match(r"=SUM\(C(\d+):I(\d+)\)(.*)", val)
            if m and int(m.group(1)) != r:
                cell_k.value = f"=SUM(C{r}:I{r}){m.group(3)}"

    # ── Atualiza Resumo Geral ─────────────────────────────────────────────────
    if "Resumo Geral" in wb_cons.sheetnames and "Contagem" in wb_cons.sheetnames:
        ws_cont = wb_cons["Contagem"]
        ws_res  = wb_cons["Resumo Geral"]

        rows_cam, rows_uv, rows_sj = [], [], []
        for i in range(1, ws_cont.max_row + 1):
            mat = ws_cont.cell(row=i, column=2).value
            if mat == "Camiseta de GV": rows_cam.append(i)
            elif mat == "Camisa U.V":   rows_uv.append(i)
            elif mat == "Short John":   rows_sj.append(i)

        def build_sum(rows, cont_col):
            return "=" + "+".join(f"Contagem!{cont_col}{r}" for r in rows)

        cont_cols_cam = ["C","D","E","F","G","H","I","J"]
        cont_cols_uv  = ["D","E","F","G","H"]

        for idx, col in enumerate(cont_cols_cam):
            ws_res.cell(row=4, column=idx+3).value = build_sum(rows_cam, col)
        for idx, col in enumerate(cont_cols_uv):
            ws_res.cell(row=7, column=idx+3).value = build_sum(rows_uv, col)
            ws_res.cell(row=10, column=idx+3).value = build_sum(rows_sj, col)

    # ── Retorna arquivo ───────────────────────────────────────────────────────
    output = io.BytesIO()
    wb_cons.save(output)
    output.seek(0)

    filename = consolidada.filename.replace(".xlsx", "_atualizado.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/health")
def health():
    return {"status": "ok"}
